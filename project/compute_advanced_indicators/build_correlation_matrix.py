"""
Build Spearman correlation matrix for bank indicators and save to Excel with heatmap.

- Numeric correlation matrix is written into Excel.
- Heatmap image is rendered in-memory (BytesIO) and embedded into Excel (no PNG file on disk).
"""

import os
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def to_dataframe(indicators: dict[str, pd.Series]) -> pd.DataFrame:
    """
    Combine series into one DataFrame by index. Drop columns with all-NaN or <2 valid points.
    """
    indicators = {k: v for k, v in indicators.items() if v is not None}
    if not indicators:
        return pd.DataFrame()

    df = pd.concat(indicators, axis=1)
    valid_cols = [c for c in df.columns if df[c].count() >= 2]
    df = df[valid_cols]
    df = df.dropna(axis=1, how="all")
    return df


def spearman_corr(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute Spearman correlation matrix.
    """
    if df.shape[1] < 2:
        return pd.DataFrame()
    return df.corr(method="spearman")


def plot_heatmap_matplotlib(
    mat: pd.DataFrame, out_png: str, title: str = "Spearman Correlation (Indicators)"
) -> BytesIO | None:
    """
    Render a heatmap and return it as an in-memory PNG (BytesIO).
    NOTE: out_png arg is kept for API compatibility but no file is written.
    """
    if mat.empty:
        return None

    data = mat.values
    labels = list(mat.columns)

    # Compact figure sizing + small fonts so it fits Excel nicely
    fig_w = max(6, len(labels) * 0.5)
    fig_h = max(5, len(labels) * 0.4)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    im = ax.imshow(data, aspect="auto")
    cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)

    ax.set_xticks(np.arange(len(labels)), labels=labels, rotation=45, ha="right")
    ax.set_yticks(np.arange(len(labels)), labels=labels)
    ax.tick_params(axis="both", labelsize=6)
    cbar.ax.tick_params(labelsize=6)
    ax.set_title(title, fontsize=10)

    # Annotate cells with correlation values
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            val = data[i, j]
            if not np.isnan(val):
                ax.text(j, i, f"{val:.2f}", ha="center", va="center", fontsize=6)

    fig.tight_layout()

    # Save to bytes instead of file
    buf = BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf


def insert_image(
    excel_path: str, sheet_name: str, img_path, top_left_cell: str = "R1"
) -> None:
    """
    Insert image into an Excel sheet.
    - If `img_path` is BytesIO or a PIL.Image, it will be embedded directly.
    - If `img_path` is a string path, it will be loaded from disk (legacy behavior).
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    if isinstance(img_path, BytesIO):
        pil_img = PILImage.open(img_path)
        img = XLImage(pil_img)
    elif hasattr(img_path, "read"):  # file-like
        pil_img = PILImage.open(img_path)
        img = XLImage(pil_img)
    elif isinstance(img_path, PILImage.Image):
        img = XLImage(img_path)
    else:
        # Fallback to path on disk if provided (kept for compatibility)
        img = XLImage(img_path)

    ws.add_image(img, top_left_cell)
    wb.save(excel_path)


def write_corr_to_excel(
    excel_path: str,
    corr: pd.DataFrame,
    sheet_name: str = "Correlation_Matrix",
) -> str:
    """
    Write correlation matrix to Excel and return cell for image placement.
    """
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        if not corr.empty:
            corr.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0)

    if not corr.empty:
        img_col = corr.shape[1] + 3  # leave a few columns gap for image
        anchor = f"{get_column_letter(img_col)}1"
    else:
        anchor = "R1"
    return anchor


def build_and_save_correlation_matrix(
    bank_data: dict,
    output_dir: str,
    filename_base: str,
) -> None:
    """
    Build only Spearman correlation matrix and save to Excel with embedded heatmap (no PNG file on disk).
    """
    indicators = bank_data["indicators_full"]
    os.makedirs(output_dir, exist_ok=True)

    # Normalize to Series in case some indicators are lists/scalars
    series_indicators: dict[str, pd.Series] = {}
    for k, v in indicators.items():
        if v is None:
            continue
        if isinstance(v, pd.Series):
            s = v
        elif isinstance(v, (list, tuple, np.ndarray)):
            s = pd.Series(v)
        else:
            s = pd.Series([v])
        s = s.dropna()
        if s.count() >= 2:
            series_indicators[k] = s

    # 1) Prepare DataFrame
    df = to_dataframe(series_indicators)

    # 2) Correlation
    corr = spearman_corr(df)

    # 3) Excel path
    excel_path = os.path.join(output_dir, f"{filename_base}_correlation.xlsx")

    # 4) Write numeric matrix
    anchor = write_corr_to_excel(excel_path, corr, sheet_name="Correlation_Matrix")

    # 5) Render heatmap to bytes and embed directly into Excel
    img_bytes = plot_heatmap_matplotlib(
        corr, out_png="", title="Spearman Correlation (Indicators)"
    )
    if img_bytes is not None:
        insert_image(excel_path, "Correlation_Matrix", img_bytes, top_left_cell=anchor)
